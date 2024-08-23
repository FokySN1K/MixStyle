import os
import shutil
import datetime
from openpyxl import load_workbook, Workbook
from typing import Optional
import pandas
import numpy as np
import re

from DataBase import config
from DataBase import models





class Work_with_Excel():

    # перед использованием класса ОБЯЗАТЕЛЬНО ставить mutex

    def __init__(self) -> None:

        # подключаемся к сессии (у двух процессов одна и та же сессия)
        self.session = config.session


        # объявляем классы для работы
        self.Product = models.Product
        self.Order = models.Order
        self.Consumer = models.Consumer
        self.Size = models.Size
        self.Color = models.Color
        self.Category = models.Category
        self.Country = models.Country

        self.Dir_Sample = "Sample"
        self.Dir_New_Product = "New_Product"
        self.Dir_New_Order = "Order"
        self.Dir_Old_Product = "Old_Product"
        self.Dir_Old_Order = "Old_Order"

        self.workbook = None


        self.categories = ['Артикул', 'Название', 'Категория' , 'Цена', 'Размер', 'Цвета', 'Компания', 'Страна', 'Количество', 'Описание']

        print("Перед использованием класса ОБЯЗАТЕЛЬНО ставить mutex")
        print("***************************", end = "\n\n")

    def Send_and_Remove_from_New_Product_to_Old_Product(self) -> Optional[str]:

        try:

            now = datetime.datetime.now()
            # переносим файл с продуктами
            product_excel_file_source = self.Dir_New_Product + "/" + os.listdir(self.Dir_New_Product)[0]
            product_excel_file_dist = self.Dir_Old_Product + "/" + "Order " + f"{now.day}_{now.month}_{now.year} {now.hour}_{now.minute}_{now.second}"
            shutil.copy2(product_excel_file_source, product_excel_file_dist)
            os.remove(product_excel_file_source)
            return None

        except Exception as E:
            print(E)
            return 'Возникла ошибка при переносе книги продуктов'

    def Set_Workbook(self, workbook: str) -> str:

        try:

            self.workbook = pandas.read_excel(
                workbook,
                dtype =
                {
                    'A': np.int64, 'B': np.str_, 'C': np.int32, 'D': np.int32,
                    'E': np.str_, 'F': np.str_, 'G':np.str_, 'H': np.int32
                }
            )

            answer: str = self.Check_Conflict_with_New_Product_and_Order()

            if answer is None:
                return 'Книга успешно загружена.'
            return answer

        except Exception as E:

            print(Exception)
            return "Не удалось загрузить книгу"

    def Check_Conflict_with_New_Product_and_Order(self) -> Optional[str]:

        ####################
        #
        #   Перед изменением БД продуктов надо проверить, что нет в заказах нет удаляемых продуктов
        #   Иначе будет пользователю нельзя будет получить текущую информацию о заказе
        #
        ##############

        ####################
        #
        #   Надо проверить:
        #
        #       1. Нет одинаковых артикулов
        #       2. В новой таблице остаются заказы, хранящиеся в БД
        #   Если есть, то возвращаем примеры, где это встречается
        #
        ##############

        try:

            datasheet = self.workbook.values
            #dataframe = pandas.DataFrame(datasheet, columns=self.categories)
            #dataframe .to_excel()

            return None

        except Exception as E:
            print(Exception)
            return "Не удалось загрузить книгу"


    def Load_to_Excel_New_Product(self):

        try:

            datasheet = self.workbook.values


            with self.session as db:

                for row in datasheet:

                    #        self.categories = ['Артикул', 'Название', 'Категория' , 'Цена',
                    #        'Размер', 'Цвета', 'Компания', 'Страна', 'Количество', 'Описание']

                    name = row[1]

                    category = re.split( ",| ", row[2])
                    category = [i for i in category if i]

                    cost = row[3]

                    article_number = row[0]

                    company = row[6]

                    describe = row[9]

                    size = row[4]

                    color = re.split( ",| ", row[5])
                    color = [i for i in color if i]

                    country = re.split(",| ", row[7])
                    country = [i for i in country if i]

                    amount = row[8]

                    Colors = []
                    for i in color:
                        Colors.append(self.Color(name=i))

                    Categories = []
                    for i in category:
                        Categories.append(self.Category(name=i))

                    Countries = []
                    for i in country:
                        Countries.append(self.Country(name=i))

                    Size = self.Size( clothing_size= size, amount=amount)

                    Product = self.Product(name=name, cost=cost,
                                           article_number=article_number,company=company,
                                           describe=describe)

                    Product.colors = Colors
                    Product.categories = Categories
                    Product.size = [Size]
                    Product.countries = Countries
                    db.merge(Product)
                    db.commit()

                #dataframe = pandas.DataFrame(datasheet, columns=self.categories)
                #dataframe .to_excel()

        except Exception as E:
            print(E)
            return 'Не удалось загрузить данные из книги в Базу Данных'


if __name__ == "__main__":


    s = Work_with_Excel()

    with config.session as db:
        models.create_tables()

    s.Set_Workbook("Sample/Product.xlsx")
    s.Load_to_Excel_New_Product()

    #with config.session as db:
        #p = db.query(models.Product).filter(models.Product.size.any(models.Size.id.in_([1]))).all()
        #print(p[0].name)
